import collections
import openpyxl

sorted_word_counts = {}
word_counts = {}

excluded_word = ['the', " ", "", "of", "to", "and", "a", "is", "am", "are", "was", "were", "i", "on", "from", "that", "if", "which", "what", "where",
 "about", "can", "could", "will", "would", "have", "has", "had", "how", "with", "it", "an", "there", "in", "for", "as", "be", "not", "any", "also", "me",
  "by", "we", "us", "do", "does", "don't", "should", "think", "from", "or", "all", "only", "but", "hong", "kong", "hk", "this", "after", "some", "more",
   "such", "like", "so", "before", "our", "may", "into", "those", "they", "these", "my", "out", "than", "much", "most", "very", "been"]


def counting_words(course):
    text_file = course + '.txt'
    with open(text_file) as fh:
        global word_counts
        global sorted_word_counts
        line = fh.readline()
        line = fh.readline()
        while line:
            comments = line.split('|')[1]
            translation_table = {ord('.'): " ", ord(','): " ", ord('?'): " ", ord('!'): " ", ord(';'): " ", ord('!'): " ", ord(':'): " "}
            comments = comments.translate(translation_table).rstrip()

            for word in comments.split(' '):
                if word.lower() in excluded_word:
                    continue
                if word_counts.get(word) is None:
                    word_counts[word] = 1
                else:
                    word_counts[word] += 1
            line = fh.readline()

def sorted_dict():
    global sorted_word_counts
    sorted_word_counts = collections.OrderedDict(reversed(sorted(word_counts.items(), key=lambda t: t[1])))

def write_to_excel(ordered_dict):
    wb = openpyxl.load_workbook('discussion.xlsx')
    sheets = wb.sheetnames
    if 'word counts' in sheets:
        del wb['word counts']
    wb.create_sheet('word counts')
    sheet = wb['word counts']

    column = 1
    for word, freq in ordered_dict.items():
        sheet.cell(column, 1).value = word
        sheet.cell(column, 2).value = freq
        column += 1
    wb.save("discussion.xlsx")
