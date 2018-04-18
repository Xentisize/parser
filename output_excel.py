import openpyxl

def write_to_excel(course):
    txt_file = course + '.txt'

    wb = openpyxl.load_workbook('discussion.xlsx')
    sheets = wb.sheetnames
    if course in sheets:
        del wb[course]
    wb.create_sheet(course)
    sheet = wb[course]

    with open(txt_file) as fh:
        column = 1
        line = fh.readline()
        while line:
            values = line.split('|')
            for index, value in enumerate(values):
                sheet.cell(column, index + 1).value = value
            column += 1
            line = fh.readline()

    wb.save("discussion.xlsx")